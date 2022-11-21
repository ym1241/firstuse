import requests
from bs4 import BeautifulSoup
from datetime import datetime as dt
from openpyxl import Workbook

now = dt.now()
targetSite = 'https://www.melon.com/chart/index.htm' #html 하나로 리스트가 안나

print("{} 현재 Melon 실시간 검색어 Top 100".format(now)) #format으로 저{}를 now로받는건가
header = {'User-agent':'Mozilla/5.0(Windows NT 6.1:WOW64;Trident/7.0;rv:11.0) Like Gecko'}
#헤더는 내가 건들것이 아님 
request = requests.get(targetSite,headers = header)
print("HTML 파일을 읽어옵니다")
print(request) #request는 멜론의 방대한 코드
html = request.text #이게 진짜 소스코드

print(html) #request 출력 몇만개

soup = BeautifulSoup(html,'html.parser')#슬라이싱

ranks = soup.findAll('div',{'class':'ellipsis rank01'}) #노래명 #div태그안에끼리

artists = soup.findAll('span',{'class':'checkEllipsis'}) #가수명 #span 태그끼리

albums = soup.findAll('div',{'class':'ellipsis rank03'}) # 앨범명

#for문으로 랭크순위 검색정보만 표시
for i in range(len(ranks)): # len 운 문자열 개수 ranks 는 무슨 문자형인지 모름
    
    artist = artists[i].text.strip().split('\n')[0] #str, split()은 문자열을 리스트로 잘라
    
    rank = ranks[i].text.strip().split('\n')[0] #str

    album = albums[i].text.strip().split('\n')[0] #앨범
    
    #print('{0:3d}.{1} - {2} - {3}'.format(i + 1, artist,rank,album))
    #print(artist) 가수명만 나옴


# 엑셀에 저장하는 부분을 작성
## 엑셀에서 저장된 리스트를 읽어와서 화면에 표시하는 부분 작성
## 읽어온 데이터를 텍스트 파일에 저장하는 부분 작성



wb= Workbook()
ws1 = wb.active #활성화 wb라는 엑셀에 ws1은 첫 시트

ws1.title = "첫번째 Sheet"
ws1["B1"] = "현재 Melon 실시간 검색어 Top"
ws1["B3"] = "순위"
ws1["C3"] = "가수이름"
ws1["D3"] = "노래제목"
ws1["E3"] = "앨범명"

print("엑셀에 저장된 정보를 가져오겠습니다.")
for t in range(len(ranks)):            #변수를 num으로 하지말아야
    ws1.cell(t+5, 2, '{}위'.format(t+1))#번호 #t는 정수 .는 문자
    t_ = ws1.cell(t+5, 2, '{}위'.format(t+1)).value
    
    artistName = artists[t].text.strip().split('\n')[0]
    ws1.cell(t +5, 3, artistName)#가수명
    t__ = ws1.cell(t +5, 3, artistName).value
    
    titleName = ranks[t].text.strip().split('\n')[0]
    ws1.cell(t +5, 4, titleName) #제목
    t___ = ws1.cell(t +5, 4, titleName).value
    
    albumName = albums[t].text.strip().split('\n')[0]
    ws1.cell(t +5, 5, albumName) #앨범
    t____ = ws1.cell(t +5, 5, albumName).value

    print("{} {}-{}\n앨범명:{}\n\n".format(t_, t__, t___, t____))


#wb.save("melontop100.xlsx") #rank를 치면 100위 제목이 나옴 artist하면 마지막 가수가 나

e = open("melontop100.txt","w")

for g in range(5,len(ranks)):
    for h in range(2,5):
    #result_Excel=""
        result_Excel = ws1.cell(row=g, column=h).value
        e.write(result_Excel)
        if h == 4:
            e.write('\n')
        elif h == 2:
            e.write(" ")
        else:
            e.write("-")
e.close()


wb.save("melontop100.xlsx")
